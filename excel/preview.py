import pandas as pd
import logging
import openpyxl

logger = logging.getLogger(__name__)

# Ranges for Projections sheet
PROJECTIONS_RANGES = {
    "Inversion": "F3:T6",
    "Inversión inicial": "F12:T26",
    "Lease": "F30:T39",
    "Detalles básicos de operación": "F46:G51",
    "P&L": "F60:T66",
    "Revenue": "F71:T84",
    "Expenses": "F86:T87",
    "GOP": "F89:T94",
    "NOI": "F97:T104",
    "Final FCF & Un-Levered IRR": "F106:T109",
    "BANK LOAN": "F183:U209"
}

def get_range_data(sheet, range_str):
    """Extracts data from an openpyxl sheet based on a range string."""
    data = []
    for row in sheet[range_str]:
        data.append([cell.value for cell in row])
    
    if not data:
        return pd.DataFrame()
        
    return pd.DataFrame(data)

def excel_to_html_preview(file_path: str) -> dict[str, str]:
    """
    Reads an Excel file and converts each relevant sheet to an HTML table.
    Specific logic for Projections ranges and FORM exclusion.
    """
    try:
        # We use openpyxl for precise range reading to avoid pandas coordinate confusion
        wb = openpyxl.load_workbook(file_path, data_only=True)
        previews = {}

        for sheet_name in wb.sheetnames:
            if sheet_name == "FORM":
                continue # Skip FORM sheet
            
            sheet = wb[sheet_name]
            
            if sheet_name == "Projections":
                html_combined = ""
                for title, range_coords in PROJECTIONS_RANGES.items():
                    df = get_range_data(sheet, range_coords)
                    if not df.empty:
                        df = df.fillna("")
                        # Identify headers (first row)
                        headers = df.iloc[0]
                        df = df[1:]
                        df.columns = headers
                        
                        html_table = df.to_html(
                            classes='preview-table', 
                            index=False, 
                            border=0,
                            justify='left'
                        )
                        html_combined += f"<div class='table-wrapper'><h4>{title}</h4>{html_table}</div>"
                previews[sheet_name] = html_combined
            else:
                # For PEM, Sizes, Listings, we take the whole sheet using pandas for convenience
                # but we'll stick to openpyxl for consistency in this function
                # Read all populated cells
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(row)
                
                df = pd.DataFrame(data)
                if not df.empty:
                    # Remove completely empty rows/cols at the end
                    df = df.dropna(how='all').dropna(axis=1, how='all')
                    df = df.fillna("")
                    
                    # Set first row as header if available
                    if len(df) > 1:
                        df.columns = df.iloc[0]
                        df = df[1:]
                    
                    html_table = df.to_html(
                        classes='preview-table', 
                        index=False, 
                        border=0,
                        justify='left'
                    )
                    previews[sheet_name] = html_table

        wb.close()
        return previews

    except Exception as e:
        logger.error(f"Failed to generate Excel preview: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return {}
